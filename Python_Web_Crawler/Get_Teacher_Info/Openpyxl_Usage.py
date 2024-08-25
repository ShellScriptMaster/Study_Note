from openpyxl import Workbook


# 实例化一个Workbook对象
# Excel_file = Workbook()
# # 启用一个excel file
# Excel_file_act = Excel_file.active
#
# # 输出excel 工作表的名称
# print(Excel_file_act.title)
# # 输出所有工作表的所有sheet 名称
# print(Excel_file.sheetnames)
#
#
# # 创建sheet
# Excel_file.create_sheet('JackySheet')
# Excel_file.create_sheet('TestingSheet')
# # 在指定位置创建sheet 从0开始
# Excel_file.create_sheet('Jessica', 1)
# print(Excel_file.sheetnames,'++++++++++++++++++++++++++++++++++')
#
# # 获取指定sheet 并输出sheet 名称
# sheet = Excel_file["JackySheet"]
# print(sheet.title)
#
# # 移动sheet 并指定偏移量
# Excel_file.move_sheet('JackySheet', -1 )
# print(Excel_file.sheetnames)
#
# Excel_file.save('./Test.xlsx')



list = [
    ['程健', '机电工程学院', 'cheng.826@hit.edu.cn', 'https://homepage.hit.edu.cn/chengjian?lang=zh'],
    ['李德友', '能源科学与工程学院', 'lideyou@hit.edu.cn', 'https://homepage.hit.edu.cn/lideyouhit?lang=zh'],
    ['宿富林', '电子与信息工程学院', 'franklin_su@hit.edu.cn', 'https://homepage.hit.edu.cn/franklinsu?lang=zh'],
    ['武高辉', '材料科学与工程学院', 'wugh@hit.edu.cn', 'https://homepage.hit.edu.cn/wugaohui?lang=zh']
]

wb = Workbook()
ws = wb.active
row = 1
ws.cell(1, 1).value = 'name'
ws.cell(2, 2).value = 'dept'
ws.cell(3, 3).value = 'mailbox'
ws.cell(4, 4).value = 'personPage'

for i in list:
    name_cell = ws.cell(row,1)
    dept_cell = ws.cell(row,2)
    mail_cell = ws.cell(row, 3)
    page_cell = ws.cell(row, 4)
    name_cell.value = i[0]
    dept_cell.value = i[1]
    mail_cell.value = i[2]
    page_cell.value = i[3]
    row += 1
    print(i,'录入成功')

wb.save('./TEST.xlsx')