import threading
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import json



url = "https://homepage.hit.edu.cn/hompage/findTeachersByName.do"
header = {
    'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36'
}

data = {
    'orderByCause':'u.modify_time desc'
}
# Home page
response = requests.post(url,header,data).text
dict_data = json.loads(response)
total_list = dict_data.get('rows')
total_info_list = []


def get_mailbox(url):

    page_RSP = requests.get(url,header).text
    soup = BeautifulSoup(page_RSP,'lxml')
    tag = soup.select('.EmailText ')
    mail_box = tag[0].text[-1::-1]
    return mail_box


def get_info(dict_data):
    Teacher_url = dict_data.get('url')
    Teacher_name = dict_data.get('userName')
    Teacher_dept = dict_data.get('department')
    common_url = 'https://homepage.hit.edu.cn/%s?lang=zh' % Teacher_url
    Teacher_mailbox = get_mailbox(common_url)
    person_info = [Teacher_name, Teacher_dept, Teacher_mailbox, common_url]
    total_info_list.append(person_info)
    print(person_info)
    return total_info_list


def multi_thread():
    threads = []
    for i in total_list:
        threads.append(
            threading.Thread(target=get_info,args=(i,))
        )
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()


multi_thread()
wb = Workbook()
ws = wb.active
row = 2
ws.cell(1, 1).value = 'name'
ws.cell(1, 2).value = 'dept'
ws.cell(1, 3).value = 'mailbox'
ws.cell(1, 4).value = 'personPage'

for i in total_info_list:
    name_cell = ws.cell(row,1)
    dept_cell = ws.cell(row,2)
    mail_cell = ws.cell(row, 3)
    page_cell = ws.cell(row, 4)
    name_cell.value = i[0]
    dept_cell.value = i[1]
    mail_cell.value = i[2]
    page_cell.value = i[3]
    row += 1
    print(i,'录入成功')

wb.save('./TEST.xlsx')



